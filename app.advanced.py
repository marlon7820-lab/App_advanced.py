# ===========================
# 8D Training App — Advanced
# Bilingual (EN/ES) + Interactive 5-Why + Fishbone + Assignments + (Optional) AI Helper
# ===========================

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import matplotlib.pyplot as plt

# ---------- Optional AI (OpenAI) ----------
AI_AVAILABLE = False
try:
    import openai
    AI_AVAILABLE = True
except Exception:
    AI_AVAILABLE = False

# ---------- Page Config & Branding ----------
st.set_page_config(
    page_title="8D Training App — Advanced",
    page_icon="📑",
    layout="wide"
)
# Hide Streamlit chrome
st.markdown("""
<style>
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 style='text-align:center'>📑 8D Training App — Advanced</h1>", unsafe_allow_html=True)

# ---------- Sidebar: Language + Mode ----------
language = st.sidebar.selectbox("Select Language / Seleccione Idioma", ["English", "Español"])
mode_training = st.sidebar.toggle("Training Mode / Modo Entrenamiento", value=True)

# ---------- Simple i18n helper ----------
T = {
    "en": {
        "report_info": "Report Information",
        "report_date": "Report Date",
        "prepared_by": "Prepared By",
        "product": "Product / Part",
        "customer": "Customer",
        "ai_helper": "AI Helper (optional)",
        "ai_about": "Get guided prompts and suggestions. Works best with concise, factual inputs.",
        "ai_key_info": "To enable AI, add your OpenAI API key in Streamlit secrets as OPENAI_API_KEY (and include 'openai' in requirements).",
        "ai_enter_key": "Enter API Key (not stored)",
        "start_coach": "Ask the Coach",
        "coach_prompt": "Describe your issue or paste your draft 5-Why here",
        "ai_not_installed": "AI package not installed. Showing heuristic helper only.",
        "ai_no_key": "No API key configured. Using heuristic helper.",
        "heuristic_title": "Heuristic Suggestions",
        "heuristic_occ": "Occurrence probing questions",
        "heuristic_det": "Detection probing questions",
        "fishbone_title": "Fishbone Diagram (Ishikawa)",
        "fishbone_note": "Enter causes by category, then click ‘Render Fishbone’.",
        "people": "People",
        "process": "Process/Method",
        "machine": "Machine/Equipment",
        "material": "Material/Components",
        "environment": "Environment",
        "measurement": "Measurement/Test",
        "add_cause": "Add another cause",
        "render_fishbone": "Render Fishbone",
        "assignments": "Assignments & Due Dates (per D-step)",
        "owner": "Owner",
        "due": "Due Date",
        "status": "Status",
        "status_opts": ["Not started", "In progress", "Done"],
        "npqp_steps": [
            ("D1: Concern Details",
             "Describe the customer concern clearly: what/where/when/how many. Include data and evidence.",
             "Ex: Customer reports intermittent static in amplifier output at end-of-line at Plant A."),
            ("D2: Similar Part Considerations",
             "Check other models/generic parts/colors/opposite hand/front-rear to learn scope.",
             "Ex: Same speaker used in Model B; compare front vs rear audio units; check opposite-hand variant."),
            ("D3: Initial Analysis",
             "Do quick checks, collect facts, and isolate the phenomenon.",
             "Ex: Visual inspect solder joints; continuity checks; connector seating; quick A/B swaps."),
            ("D4: Implement Containment",
             "Temporary measures to protect customer while you find root cause.",
             "Ex: 100% inspection, quarantine lot, temporary shielding, extra test step."),
            ("D5: Final Analysis",
             "Use 5-Why: Occurrence (why it happened) and Detection (why it escaped). Add more Whys if needed.",
             ""),
            ("D6: Permanent Corrective Actions",
             "Define actions that eliminate the root cause and prevent recurrence.",
             "Ex: Update solder profile, retrain, revise WI, add automated inspection."),
            ("D7: Countermeasure Confirmation",
             "Verify actions are effective over time.",
             "Ex: Run verification builds, life tests, capability studies (CPK), monitor early field data."),
            ("D8: Follow-up Activities (Lessons Learned / Recurrence Prevention)",
             "Standardize changes, update PFMEA/Control Plan/SOP, communicate lessons learned.",
             "Ex: Update PFMEA, add control plan checks, train team, add audit points.")
        ],
        "d5_training": (
            "**Training Guidance:** Use 5-Why to reach process-level causes.\n\n"
            "**Occurrence Example (5-Whys):**\n"
            "1. Cold solder joint on DSP chip\n"
            "2. Solder temp too low\n"
            "3. Operator didn’t follow profile\n"
            "4. Work instructions unclear\n"
            "5. No visual confirmation step\n\n"
            "**Detection Example (5-Whys):**\n"
            "1. QA missed cold joint\n"
            "2. Checklist incomplete\n"
            "3. No automated vision/test\n"
            "4. Batch testing not performed\n"
            "5. Early warning trend not tracked\n\n"
            "**Root Cause Example:**\n"
            "Insufficient solder process control + inadequate QA checklist allowed defect to escape."
        ),
        "occurrence": "Occurrence Analysis",
        "detection": "Detection Analysis",
        "why": "Why",
        "add_why": "Add another Why",
        "root_cause": "Root Cause (summary after 5-Whys)",
        "your_answer": "Your Answer",
        "save": "Save 8D Report",
        "saved": "8D Report saved successfully.",
        "download": "Download XLSX",
        "no_answers": "No answers yet. Please complete some fields before saving."
    },
    "es": {
        "report_info": "Información del Reporte",
        "report_date": "Fecha del Reporte",
        "prepared_by": "Preparado por",
        "product": "Producto / Parte",
        "customer": "Cliente",
        "ai_helper": "Asistente IA (opcional)",
        "ai_about": "Obtenga preguntas guía y sugerencias. Funciona mejor con entradas concisas y objetivas.",
        "ai_key_info": "Para habilitar IA, agregue su clave de OpenAI en secretos de Streamlit como OPENAI_API_KEY (y añada 'openai' a requirements).",
        "ai_enter_key": "Ingrese la clave (no se guarda)",
        "start_coach": "Preguntar al Asistente",
        "coach_prompt": "Describa el problema o pegue su borrador 5-Why aquí",
        "ai_not_installed": "Paquete de IA no instalado. Mostrando asistente heurístico.",
        "ai_no_key": "No hay clave de API configurada. Usando asistente heurístico.",
        "heuristic_title": "Sugerencias Heurísticas",
        "heuristic_occ": "Preguntas para profundizar (Ocurrencia)",
        "heuristic_det": "Preguntas para profundizar (Detección)",
        "fishbone_title": "Diagrama de Espina de Pescado (Ishikawa)",
        "fishbone_note": "Ingrese causas por categoría y luego haga clic en ‘Generar Diagrama’.",
        "people": "Personas",
        "process": "Proceso/Método",
        "machine": "Máquina/Equipo",
        "material": "Material/Componentes",
        "environment": "Entorno",
        "measurement": "Medición/Prueba",
        "add_cause": "Agregar otra causa",
        "render_fishbone": "Generar Diagrama",
        "assignments": "Responsables y Fechas (por paso D)",
        "owner": "Responsable",
        "due": "Fecha Límite",
        "status": "Estado",
        "status_opts": ["No iniciado", "En progreso", "Terminado"],
        "npqp_steps": [
            ("D1: Detalles de la Queja",
             "Describa claramente la queja: qué/dónde/cuándo/cuántos. Incluya datos y evidencias.",
             "Ej: Cliente reporta estática intermitente en el amplificador en fin de línea en Planta A."),
            ("D2: Consideración de Partes Similares",
             "Revise otros modelos/partes genéricas/colores/mano opuesta/delantero-trasero para entender el alcance.",
             "Ej: El mismo parlante usado en Modelo B; comparar delantero vs trasero; validar mano opuesta."),
            ("D3: Análisis Inicial",
             "Realice verificaciones rápidas, recolecte hechos y aísle el fenómeno.",
             "Ej: Inspección visual de soldaduras; pruebas de continuidad; asientos de conectores; swaps A/B."),
            ("D4: Implementar Contención",
             "Medidas temporales para proteger al cliente mientras encuentra la causa raíz.",
             "Ej: Inspección 100%, poner lote en cuarentena, blindaje temporal, paso de prueba adicional."),
            ("D5: Análisis Final",
             "Use 5-Why: Ocurrencia (por qué ocurrió) y Detección (por qué escapó). Agregue más ‘porqués’ si necesita.",
             ""),
            ("D6: Acciones Correctivas Permanentes",
             "Defina acciones que eliminen la causa raíz y eviten recurrencia.",
             "Ej: Actualizar perfil de soldadura, reentrenar, revisar instrucciones, añadir inspección automática."),
            ("D7: Confirmación de Contramedidas",
             "Verifique que las acciones sean efectivas en el tiempo.",
             "Ej: Lotes de verificación, pruebas de vida, estudios de capacidad (CPK), monitoreo temprana de campo."),
            ("D8: Seguimiento (Lecciones Aprendidas / Prevención de Recurrencia)",
             "Estandarice cambios, actualice PFMEA/Plan de Control/POE, comunique lecciones aprendidas.",
             "Ej: Actualizar PFMEA, sumar controles al plan, entrenar equipo, añadir puntos de auditoría.")
        ],
        "d5_training": (
            "**Guía de Entrenamiento:** Use 5-Why para llegar a causas a nivel de proceso.\n\n"
            "**Ejemplo Ocurrencia (5-Why):**\n"
            "1. Unión fría de soldadura en DSP\n"
            "2. Temperatura de soldado baja\n"
            "3. Operador no siguió el perfil\n"
            "4. Instrucciones poco claras\n"
            "5. Sin verificación visual\n\n"
            "**Ejemplo Detección (5-Why):**\n"
            "1. QA no detectó la unión fría\n"
            "2. Checklist incompleto\n"
            "3. Sin visión/prueba automática\n"
            "4. No se realizaron pruebas por lote\n"
            "5. No se monitoreó la señal temprana\n\n"
            "**Ejemplo Causa Raíz:**\n"
            "Control insuficiente del proceso de soldadura + checklist de QA inadecuado permitió fuga del defecto."
        ),
        "occurrence": "Análisis de Ocurrencia",
        "detection": "Análisis de Detección",
        "why": "¿Por qué?",
        "add_why": "Agregar otro ¿Por qué?",
        "root_cause": "Causa Raíz (resumen tras 5-Why)",
        "your_answer": "Su Respuesta",
        "save": "Guardar Reporte 8D",
        "saved": "Reporte 8D guardado con éxito.",
        "download": "Descargar XLSX",
        "no_answers": "Aún no hay respuestas. Complete algunos campos antes de guardar."
    }
}
L = T["en"] if language == "English" else T["es"]

# ---------- Report Info ----------
st.subheader(L["report_info"])
today_str = datetime.datetime.today().strftime("%B %d, %Y") if language == "English" else datetime.datetime.today().strftime("%d/%m/%Y")
col_a, col_b, col_c, col_d = st.columns(4)
with col_a:
    report_date = st.text_input(f"📅 {L['report_date']}", value=today_str, key="rp_date")
with col_b:
    prepared_by = st.text_input(f"✍️ {L['prepared_by']}", key="rp_by")
with col_c:
    product = st.text_input(f"🔧 {L['product']}", key="rp_prod")
with col_d:
    customer = st.text_input(f"👤 {L['customer']}", key="rp_cust")

# ---------- NPQP Steps + State ----------
npqp_steps = L["npqp_steps"]  # list of tuples (step, note, example)
if "answers" not in st.session_state:
    st.session_state.answers = {step: "" for step, _, _ in npqp_steps}
if "owners" not in st.session_state:
    st.session_state.owners = {step: "" for step, _, _ in npqp_steps}
if "dues" not in st.session_state:
    st.session_state.dues = {step: None for step, _, _ in npqp_steps}
if "status" not in st.session_state:
    st.session_state.status = {step: L["status_opts"][0] for step, _, _ in npqp_steps}

# D5-specific session state
st.session_state.setdefault("d5_occ", [""] * 5)
st.session_state.setdefault("d5_det", [""] * 5)
st.session_state.setdefault("d5_root", "")

# ---------- AI Helper (optional) ----------
with st.expander(f"🤖 {L['ai_helper']}", expanded=False):
    st.caption(L["ai_about"])
    if not AI_AVAILABLE:
        st.warning(L["ai_not_installed"])
    api_key = st.text_input(f"🔐 {L['ai_enter_key']}", type="password")
    if AI_AVAILABLE:
        if not api_key and "OPENAI_API_KEY" in st.secrets:
            api_key = st.secrets["OPENAI_API_KEY"]
        if api_key:
            openai.api_key = api_key
        else:
            st.info(L["ai_no_key"])

    user_issue = st.text_area(f"💬 {L['coach_prompt']}", height=130)
    if st.button(f"🚀 {L['start_coach']}"):
        suggestions = []
        # Heuristic helper (always available)
        occ_q = [
            "What changed (materials, machine, method, environment) right before the first failure?",
            "Can you reproduce the defect reliably? Under what conditions?",
            "Is the problem isolated to certain lots, lines, shifts, or suppliers?",
            "Which CTQ or spec is violated? What data trend shows the shift?",
            "What evidence rules out common red herrings?"
        ]
        det_q = [
            "Which check should have detected it (test, visual, measurement)?",
            "Is the control plan/PFMEA aligned with actual risks?",
            "Was the inspection capable (Gage R&R, sensitivity)?",
            "Were criteria or sampling insufficient for this failure mode?",
            "Any prior warnings, near-misses, or customer escapes?"
        ]
        if language == "Español":
            occ_q = [
                "¿Qué cambió (materiales, máquina, método, entorno) justo antes del primer fallo?",
                "¿Puede reproducirse el defecto de forma confiable? ¿Bajo qué condiciones?",
                "¿Se limita a ciertos lotes, líneas, turnos o proveedores?",
                "¿Qué CTQ o especificación se viola? ¿Qué tendencia de datos muestra el cambio?",
                "¿Qué evidencias descartan falsos indicios comunes?"
            ]
            det_q = [
                "¿Qué control debía detectarlo (prueba, visual, medición)?",
                "¿El plan de control/PFMEA refleja los riesgos reales?",
                "¿La inspección era capaz (R&R, sensibilidad)?",
                "¿Criterios o muestreo insuficientes para este modo de falla?",
                "¿Hubo alertas previas, casi-incidentes o fugas al cliente?"
            ]
        suggestions.append(("Occurrence/Ocurrencia", occ_q))
        suggestions.append(("Detection/Detección", det_q))

        if AI_AVAILABLE and api_key:
            try:
                prompt = f"""You are a quality problem-solving coach for electronics (radios, speakers, amplifiers).
User note: {user_issue}
Return: (1) clarifying questions, (2) likely categories (People/Process/Machine/Material/Environment/Measurement),
(3) example Occurrence and Detection Why-questions, (4) one possible root cause pattern to investigate.
Be concise, bullet each section, write in {language}."""
                # Use Responses API if available; fallback to ChatCompletion if older
                try:
                    resp = openai.ChatCompletion.create(
                        model="gpt-4o-mini",
                        messages=[{"role": "system", "content": "You are a concise quality coach."},
                                  {"role": "user", "content": prompt}],
                        temperature=0.3
                    )
                    ai_text = resp["choices"][0]["message"]["content"]
                except Exception:
                    # Minimal fallback
                    ai_text = ""
                if ai_text:
                    st.success("AI Coach Suggestions:")
                    st.write(ai_text)
            except Exception as e:
                st.warning(f"AI error: {e}")

        st.markdown(f"### 🧭 {L['heuristic_title']}")
        st.markdown(f"**• {L['heuristic_occ']}:**")
        for q in occ_q: st.write(f"- {q}")
        st.markdown(f"**• {L['heuristic_det']}:**")
        for q in det_q: st.write(f"- {q}")

# ---------- Tabs for D1–D8 ----------
tabs = st.tabs([s for s, _, _ in npqp_steps])

for i, (step, note, example) in enumerate(npqp_steps):
    with tabs[i]:
        st.markdown(f"### {step}")

        if mode_training:
            if step.startswith("D5"):
                st.info(L["d5_training"])
            else:
                st.info(f"**Training Guidance / Guía:** {note}\n\n💡 **Example / Ejemplo:** {example}")

        # Assignments
        st.markdown(f"**{L['assignments']}**")
        ca, cb, cc = st.columns([2, 1.5, 1.5])
        with ca:
            st.session_state.owners[step] = st.text_input(f"👤 {L['owner']} — {step}", value=st.session_state.owners[step], key=f"own_{i}")
        with cb:
            st.session_state.dues[step] = st.date_input(f"📅 {L['due']} — {step}", value=st.session_state.dues[step], key=f"due_{i}")
        with cc:
            st.session_state.status[step] = st.selectbox(f"📌 {L['status']} — {step}", options=L["status_opts"], index=L["status_opts"].index(st.session_state.status[step]) if st.session_state.status[step] in L["status_opts"] else 0, key=f"st_{i}")

        st.markdown("---")

        if step.startswith("D5"):
            # Occurrence
            st.subheader(L["occurrence"])
            for idx, val in enumerate(st.session_state.d5_occ):
                st.session_state.d5_occ[idx] = st.text_input(f"{L['why']} #{idx+1} — {L['occurrence']}", value=val, key=f"d5o_{idx}")
            if st.button(f"➕ {L['add_why']} — {L['occurrence']}", key="add_occ"):
                st.session_state.d5_occ.append("")

            # Detection
            st.subheader(L["detection"])
            for idx, val in enumerate(st.session_state.d5_det):
                st.session_state.d5_det[idx] = st.text_input(f"{L['why']} #{idx+1} — {L['detection']}", value=val, key=f"d5d_{idx}")
            if st.button(f"➕ {L['add_why']} — {L['detection']}", key="add_det"):
                st.session_state.d5_det.append("")

            st.session_state.d5_root = st.text_area(L["root_cause"], value=st.session_state.d5_root, height=120, key="d5root")

            # Compose D5 answer (for Excel)
            occ_txt = "\n".join([w for w in st.session_state.d5_occ if w.strip()])
            det_txt = "\n".join([w for w in st.session_state.d5_det if w.strip()])
            st.session_state.answers[step] = f"{L['occurrence']}:\n{occ_txt}\n\n{L['detection']}:\n{det_txt}"
        else:
            st.session_state.answers[step] = st.text_area(f"📝 {L['your_answer']} — {step}", value=st.session_state.answers[step], height=160, key=f"ans_{i}")

# ---------- Fishbone Diagram ----------
st.markdown("---")
st.header(f"🐟 {L['fishbone_title']}")
st.caption(L["fishbone_note"])

if "fishbone" not in st.session_state:
    st.session_state.fishbone = {
        "People": [""],
        "Process/Method": [""],
        "Machine/Equipment": [""],
        "Material/Components": [""],
        "Environment": [""],
        "Measurement/Test": [""],
    }

# localized category keys -> internal keys
cat_map = {
    L["people"]: "People",
    L["process"]: "Process/Method",
    L["machine"]: "Machine/Equipment",
    L["material"]: "Material/Components",
    L["environment"]: "Environment",
    L["measurement"]: "Measurement/Test",
}

cols = st.columns(3)
cats_local = list(cat_map.keys())
for idx, cat_local in enumerate(cats_local):
    with cols[idx % 3]:
        st.markdown(f"**{cat_local}**")
        key_internal = cat_map[cat_local]
        entries = st.session_state.fishbone.get(key_internal, [""])
        # render inputs
        for j in range(len(entries)):
            entries[j] = st.text_input(f"{cat_local} cause #{j+1}", value=entries[j], key=f"fb_{key_internal}_{j}")
        if st.button(f"➕ {L['add_cause']} — {cat_local}", key=f"add_{key_internal}"):
            entries.append("")
        st.session_state.fishbone[key_internal] = entries

if st.button(f"📈 {L['render_fishbone']}"):
    # Simple text-based fishbone plotted as categories radial list
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.axis('off')
    center_text = "Problem" if language == "English" else "Problema"
    ax.text(0.5, 0.5, center_text, ha='center', va='center', fontsize=16, fontweight='bold')
    # Place categories around
    positions = [(0.15, 0.8), (0.85, 0.8), (0.15, 0.2), (0.85, 0.2), (0.15, 0.5), (0.85, 0.5)]
    for (cat_local, pos) in zip(cats_local, positions):
        key_internal = cat_map[cat_local]
        ax.text(pos[0], pos[1], cat_local, fontsize=12, fontweight='bold', ha='center')
        causes = [c for c in st.session_state.fishbone.get(key_internal, []) if c.strip()]
        # list causes under the category
        y = pos[1] - 0.06
        for c in causes[:6]:
            ax.text(pos[0], y, f"• {c}", fontsize=10, ha='center', va='top')
            y -= 0.045
    st.pyplot(fig)

# ---------- Save to Excel ----------
step_colors = {
    "D1: Concern Details": "ADD8E6",
    "D2: Similar Part Considerations": "90EE90",
    "D3: Initial Analysis": "FFFF99",
    "D4: Implement Containment": "FFD580",
    "D5: Final Analysis": "FF9999",
    "D6: Permanent Corrective Actions": "D8BFD8",
    "D7: Countermeasure Confirmation": "E0FFFF",
    "D8: Follow-up Activities (Lessons Learned / Recurrence Prevention)": "D3D3D3",
    "D1: Detalles de la Queja": "ADD8E6",
    "D2: Consideración de Partes Similares": "90EE90",
    "D3: Análisis Inicial": "FFFF99",
    "D4: Implementar Contención": "FFD580",
    "D5: Análisis Final": "FF9999",
    "D6: Acciones Correctivas Permanentes": "D8BFD8",
    "D7: Confirmación de Contramedidas": "E0FFFF",
    "D8: Seguimiento (Lecciones Aprendidas / Prevención de Recurrencia)": "D3D3D3"
}

if st.button(f"💾 {L['save']}"):
    data_rows = [(s, st.session_state.answers[s],
                  st.session_state.d5_root if s.startswith("D5") else "",
                  st.session_state.owners[s], st.session_state.dues[s], st.session_state.status[s])
                 for s, _, _ in npqp_steps]

    if not any(ans for _, ans, _, _, _, _ in data_rows):
        st.error(f"⚠️ {L['no_answers']}")
    else:
        xlsx_file = "NPQP_8D_Advanced.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "NPQP 8D Report"

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = "Nissan NPQP 8D Report"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # Report info
        ws["A3"] = "Report Date / Fecha"
        ws["B3"] = report_date
        ws["A4"] = "Prepared By / Preparado por"
        ws["B4"] = prepared_by
        ws["A5"] = "Product / Part"
        ws["B5"] = product
        ws["A6"] = "Customer / Cliente"
        ws["B6"] = customer

        # Headers
        headers = ["Step", "Your Answer", "Root Cause", "Owner", "Due Date", "Status"]
        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        row = 8
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = header_fill

        # Content
        row = 9
        for step, ans, root, owner, due, status in data_rows:
            ws.cell(row=row, column=1, value=step)
            ws.cell(row=row, column=2, value=ans)
            ws.cell(row=row, column=3, value=root)
            ws.cell(row=row, column=4, value=owner)
            ws.cell(row=row, column=5, value=str(due) if due else "")
            ws.cell(row=row, column=6, value=status)

            fill_color = step_colors.get(step, "FFFFFF")
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

        # Column widths
        widths = [28, 46, 40, 20, 18, 18]
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w

        wb.save(xlsx_file)
        st.success(f"✅ {L['saved']}")
        with open(xlsx_file, "rb") as f:
            st.download_button(f"📥 {L['download']}", f, file_name=xlsx_file)
